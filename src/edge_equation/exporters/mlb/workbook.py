"""
Multi-tab MLB daily workbook (mlb_daily.xlsx).

Tabs (in order):
    Today's Card   — combined picks-of-the-day with brand header
    Moneyline      — every ML pick the gate let through
    Run Line       — every RL pick (we bet the +1.5 underdog by default)
    Totals         — full-game over/under
    First 5        — F5 ML / F5 RL / F5 Total stacked
    First Inning   — NRFI / YRFI from v1's NRFI engine
    Team Totals    — per-team scoring lines
    Backtest       — rolling per-market ROI / Brier / hit_rate

Every per-market tab carries the same row schema so the in-cell grade
formula (D → A+) and the Kelly tier formula resolve uniformly.

The grade / Kelly formulas are pure cell formulas — no macros — so the
workbook opens cleanly in Numbers iOS / Excel / LibreOffice and
recomputes live as the daily run rewrites the rows.

Brand stamp: row 1 of "Today's Card" carries the BRAND_TAGLINE constant
("Facts. Not Feelings."). The compliance sanitizer is responsible for
stripping edge / kelly columns from any *public* CSV mirror; the xlsx
itself is a private operator artifact.
"""
from __future__ import annotations

from pathlib import Path
from typing import Iterable

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName


BRAND_TAGLINE = "Facts. Not Feelings."

HEADER_FILL = PatternFill("solid", fgColor="1E2229")
HEADER_FONT = Font(name="Helvetica Neue", size=12, bold=True, color="5BC0E8")
DATA_FONT = Font(name="Helvetica Neue", size=11)
BRAND_FONT = Font(name="Helvetica Neue", size=14, bold=True, italic=True, color="5BC0E8")


PICK_COLUMNS = [
    ("date", 12),
    ("matchup", 18),
    ("bet_type", 12),
    ("pick", 22),
    ("line", 8),
    ("model_prob", 11),
    ("market_prob", 11),
    ("edge_pct", 10),
    ("decimal_odds", 12),
    ("american_odds", 12),
    ("kelly_pct", 10),
    ("kelly_advice", 12),
    ("grade", 8),
    ("book", 14),
    ("game_time", 18),
]

# Column letters for formula references — derived from PICK_COLUMNS order.
COL = {name: get_column_letter(i + 1) for i, (name, _) in enumerate(PICK_COLUMNS)}


# Grade formula: ladder over edge_pct (column COL["edge_pct"]).
# Mirrors v1 src/edge_equation/math/scoring.py grade thresholds, restated
# in percent because the workbook stores edge as percent (e.g. 4.5) not
# fraction (0.045). If scoring.py thresholds drift, this string follows.
def _grade_formula(row: int) -> str:
    e = f"{COL['edge_pct']}{row}"
    return (
        f'=IF({e}="","",'
        f'IF({e}>=8,"A+",'
        f'IF({e}>=5,"A",'
        f'IF({e}>=3,"B",'
        f'IF({e}>=0,"C",'
        f'IF({e}>=-3,"D","F")))))) '
    ).strip()


def _kelly_advice_formula(row: int) -> str:
    """In-cell mirror of kelly._tier(). Uses kelly_pct column."""
    k = f"{COL['kelly_pct']}{row}"
    return (
        f'=IF({k}="","PASS",'
        f'IF({k}<=0.5,"PASS",'
        f'IF({k}<=1.5,"0.5u",'
        f'IF({k}<=3,"1u",'
        f'IF({k}<=5,"2u","3u"))))) '
    ).strip()


def _write_pick_header(ws):
    for col_idx, (name, width) in enumerate(PICK_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=name)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="left", vertical="center")
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.freeze_panes = "A2"


def _write_pick_row(ws, row_idx: int, pick: dict):
    """Write a single pick row. Grade + kelly_advice cells are formulas
    so they recompute if numeric inputs are edited by hand."""
    values = {
        "date": pick.get("date"),
        "matchup": pick.get("matchup"),
        "bet_type": pick.get("bet_type"),
        "pick": pick.get("pick"),
        "line": pick.get("line"),
        "model_prob": pick.get("model_prob"),
        "market_prob": pick.get("market_prob"),
        "edge_pct": pick.get("edge_pct"),
        "decimal_odds": pick.get("decimal_odds"),
        "american_odds": pick.get("american_odds"),
        "kelly_pct": pick.get("kelly_pct"),
        "book": pick.get("book"),
        "game_time": pick.get("game_time"),
    }
    for name, value in values.items():
        ws[f"{COL[name]}{row_idx}"] = value
        ws[f"{COL[name]}{row_idx}"].font = DATA_FONT
    # Formula cells last so they reference the values written above.
    ws[f"{COL['kelly_advice']}{row_idx}"] = _kelly_advice_formula(row_idx)
    ws[f"{COL['grade']}{row_idx}"] = _grade_formula(row_idx)
    ws[f"{COL['model_prob']}{row_idx}"].number_format = "0.000"
    ws[f"{COL['market_prob']}{row_idx}"].number_format = "0.000"
    ws[f"{COL['edge_pct']}{row_idx}"].number_format = "0.00"
    ws[f"{COL['kelly_pct']}{row_idx}"].number_format = "0.00"
    ws[f"{COL['decimal_odds']}{row_idx}"].number_format = "0.000"


def _write_market_tab(wb: Workbook, title: str, picks: Iterable[dict]):
    ws = wb.create_sheet(title)
    _write_pick_header(ws)
    for offset, pick in enumerate(picks, start=2):
        _write_pick_row(ws, offset, pick)


def _write_todays_card(wb: Workbook, picks: Iterable[dict], as_of: str | None):
    ws = wb.active
    ws.title = "Today's Card"
    ws["A1"] = BRAND_TAGLINE
    ws["A1"].font = BRAND_FONT
    ws.merge_cells("A1:E1")
    if as_of:
        ws["F1"] = f"as of {as_of}"
        ws["F1"].font = Font(italic=True, color="8A8A7A", size=10)
    # Headers shift to row 3, picks to row 4+.
    for col_idx, (name, width) in enumerate(PICK_COLUMNS, start=1):
        cell = ws.cell(row=3, column=col_idx, value=name)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.freeze_panes = "A4"
    for offset, pick in enumerate(picks, start=4):
        _write_pick_row(ws, offset, pick)


def _write_backtest_tab(wb: Workbook, summary_by_bet_type: list[dict]):
    ws = wb.create_sheet("Backtest")
    cols = [
        ("bet_type", 14), ("bets", 8), ("wins", 8), ("losses", 8),
        ("pushes", 8), ("hit_rate", 10), ("roi_pct", 10),
        ("brier", 10), ("clv_pct", 10), ("status", 10),
    ]
    for col_idx, (name, width) in enumerate(cols, start=1):
        cell = ws.cell(row=1, column=col_idx, value=name)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.freeze_panes = "A2"
    for row_idx, row in enumerate(summary_by_bet_type or [], start=2):
        for col_idx, (name, _) in enumerate(cols, start=1):
            ws.cell(row=row_idx, column=col_idx, value=row.get(name))


# Public market-tab name registry. Order = order tabs appear in the workbook.
MARKET_TABS = [
    ("Moneyline",    "moneyline"),
    ("Run Line",     "run_line"),
    ("Totals",       "totals"),
    ("First 5",      "first_5"),
    ("First Inning", "first_inning"),
    ("Team Totals",  "team_totals"),
]


def build_workbook(
    todays_card: list[dict],
    picks_by_market: dict[str, list[dict]],
    backtest_summary: list[dict],
    as_of: str | None = None,
) -> Workbook:
    """Assemble the full multi-tab workbook from already-gated picks.

    Args:
        todays_card: filtered top picks for the Today's Card sheet.
        picks_by_market: {bet_type: [pick_dict, ...]} for each market tab.
            Empty markets get an empty tab (header only) so the workbook
            shape is stable across days — important for users who built
            iOS Shortcuts pointing at fixed cell ranges.
        backtest_summary: rows for the Backtest tab.
        as_of: timestamp string for the Today's Card subheader.
    """
    wb = Workbook()
    _write_todays_card(wb, todays_card, as_of)
    for tab_title, bet_type in MARKET_TABS:
        _write_market_tab(wb, tab_title, picks_by_market.get(bet_type, []))
    _write_backtest_tab(wb, backtest_summary)

    # payout_mult defined name preserved from the legacy props workbook
    # so any iOS Shortcut keyed off it keeps working.
    wb.defined_names["payout_mult"] = DefinedName(
        "payout_mult", attr_text="'Today''s Card'!$A$1"
    )
    return wb


def write_workbook(
    out_path: Path,
    todays_card: list[dict],
    picks_by_market: dict[str, list[dict]],
    backtest_summary: list[dict],
    as_of: str | None = None,
) -> Path:
    out_path = Path(out_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb = build_workbook(todays_card, picks_by_market, backtest_summary, as_of=as_of)
    wb.save(out_path)
    return out_path
