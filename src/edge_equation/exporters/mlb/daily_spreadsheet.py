"""
MLB Daily Spreadsheet Exporter
==============================
Produces a 6-tab spreadsheet of game-results bets (Moneyline, Run Line,
Totals, First 5, First Inning, Team Totals) covering season-to-date
backfill plus projections for today's slate. Outputs land in
public/data/mlb/ so a Vercel-hosted frontend can serve them.

Usage:
    python -m exporters.mlb.daily_spreadsheet
    python -m exporters.mlb.daily_spreadsheet --date 2026-05-02
    python -m exporters.mlb.daily_spreadsheet --no-push
    python -m exporters.mlb.daily_spreadsheet --season 2026 --output-dir public/data/mlb
"""

from __future__ import annotations

import argparse
import csv
import json
import os
import subprocess
import sys
from datetime import datetime, timedelta
from pathlib import Path
from typing import Iterable

try:
    from zoneinfo import ZoneInfo
except ImportError:
    ZoneInfo = None

import requests

# In v1's src-layout this file is src/edge_equation/exporters/mlb/daily_spreadsheet.py
# so REPO_ROOT is parents[4] (mlb -> exporters -> edge_equation -> src -> root).
REPO_ROOT = Path(__file__).resolve().parents[4]

# Feature flag — verbatim orchestrator no-ops unless explicitly enabled.
# Set EDGE_FEATURE_SPREADSHEET_PIPELINE=on (env) to run. The cron workflow
# sets it; local dev runs default off so an accidental run doesn't pollute
# public/data/mlb/.
FEATURE_FLAG_ENV = "EDGE_FEATURE_SPREADSHEET_PIPELINE"


def _feature_flag_on() -> bool:
    return os.environ.get(FEATURE_FLAG_ENV, "").strip().lower() in (
        "1", "on", "true", "yes",
    )


from edge_equation.scrapers.mlb.mlb_game_scraper import MLBGameScraper, TEAM_MAP
from edge_equation.scrapers.mlb.mlb_odds_scraper import MLBOddsScraper
from edge_equation.scrapers.mlb.mlb_pitcher_scraper import MLBPitcherScraper
from edge_equation.scrapers.mlb.mlb_weather_scraper import MLBWeatherScraper
from edge_equation.scrapers.mlb.mlb_lineup_scraper import MLBLineupScraper
from edge_equation.exporters.mlb.projections import (
    ProjectionModel,
    prob_over,
    prob_over_under_poisson,
    prob_over_under_smart,
    TOTAL_SD,
    TEAM_TOTAL_SD,
    F5_TOTAL_SD,
)
from edge_equation.exporters.mlb.kelly import (
    kelly_advice, edge_pct, tier_from_pct, DEFAULT_DECIMAL_ODDS,
)
from edge_equation.exporters.mlb.backtest import BacktestEngine
from edge_equation.exporters.mlb.clv_tracker import ClvTracker
from edge_equation.exporters.mlb.splits_loader import SplitsLoader


SEASON_DEFAULT = 2026
SEASON_OPENING_DAY = "{season}-03-20"
DEFAULT_OUTPUT_DIR = REPO_ROOT / "public" / "data" / "mlb"
TOTAL_LINES = (8.5, 9.0, 9.5)
TEAM_TOTAL_LINES = (3.5, 4.5)

# Today's Card filtering — the user's strategy is to play only the
# highest-edge plays. Edge thresholds are tuned per market because
# variance and typical vig differ:
#   - Moneyline: heavy favorites (-200+) have huge price-induced variance,
#     so demand a meatier edge.
#   - Run line / Team totals: moderate variance, moderate vig (~-115).
#   - Totals / First 5: low vig (-110/-110), can act on tighter edges.
#   - First inning (NRFI/YRFI): binary but the market is less efficient,
#     so we want at least a 4% edge before believing it.
DEFAULT_MIN_EDGE_PCT = 3.0
DEFAULT_TOP_N = 10

# Market gating — per BRAND_GUIDE Product Rules: "A market is included
# on the daily card only when its rolling 200+ bet backtest shows ≥+1%
# ROI AND Brier <0.246." Markets that fail still appear in their own
# tabs (transparency) but stay off the headline card.
DEFAULT_MIN_GATE_BETS = 200
DEFAULT_MIN_GATE_ROI = 1.0
DEFAULT_MAX_GATE_BRIER = 0.246
DEFAULT_EDGE_THRESHOLDS_BY_MARKET = {
    "moneyline":     4.0,
    "run_line":      3.0,
    "totals":        2.5,
    "first_5":       2.5,
    "first_inning":  4.0,
    "team_totals":   3.0,
}

# Portfolio sizing — sum of half-Kelly% across all bets on a single
# game is capped at this value (% of bankroll). Same-game bets are
# correlated (an OVER total + a fav ML are both juiced by an offensive
# eruption), so full-Kelly across all of them over-stakes the slate.
DEFAULT_PORTFOLIO_CAP_PER_GAME = 6.0


def _today_et() -> str:
    if ZoneInfo is not None:
        return datetime.now(ZoneInfo("America/New_York")).strftime("%Y-%m-%d")
    return datetime.utcnow().strftime("%Y-%m-%d")


def _ou_label(actual: float, line: float) -> str:
    if actual > line:
        return "OVER"
    if actual < line:
        return "UNDER"
    return "PUSH"


def _best_total_kelly(mean: float, lines: tuple[float, ...], sd: float) -> dict:
    """Pick the (line, side) with the highest half-Kelly at the default price.

    Uses NegBin when sd implies over-dispersion (typical for MLB run
    totals: empirical variance > mean) and Poisson otherwise. Half-point
    lines have P(push)=0; whole-number lines correctly isolate the push
    probability.
    """
    best = None
    for line in lines:
        p_over, p_under, _push = prob_over_under_smart(line, mean, sd)
        for side, prob in (("OVER", p_over), ("UNDER", p_under)):
            adv = kelly_advice(prob)
            if best is None or adv["kelly_pct"] > best["kelly_pct"]:
                best = {**adv, "line": line, "side": side}
    return best


def _market_total_kelly(
    mean: float,
    sd: float,
    market_totals: list[dict],
) -> dict | None:
    """Best Kelly across whichever totals lines the book is actually offering.

    NegBin-when-over-dispersed via prob_over_under_smart so probabilities
    aren't artificially pumped by the equidispersion assumption.
    Optimizes half-Kelly fraction; returns the (line, side) combo with
    the highest edge plus its edge_pct.
    """
    best = None
    for offer in market_totals:
        line = offer.get("point")
        if line is None:
            continue
        p_over, p_under, _push = prob_over_under_smart(line, mean, sd)
        for side_key, side_label, prob in (
            ("over", "OVER", p_over),
            ("under", "UNDER", p_under),
        ):
            price = offer.get(side_key)
            if not price:
                continue
            adv = kelly_advice(prob, decimal_odds=price["decimal"])
            cand = {
                **adv,
                "line": line,
                "side": side_label,
                "market_decimal": price["decimal"],
                "market_american": price["american"],
                "book": price["book"],
                "edge_pct": edge_pct(prob, price["decimal"]),
            }
            if best is None or cand["kelly_pct"] > best["kelly_pct"]:
                best = cand
    return best


def _ml_market_price(market: dict, side: str) -> dict | None:
    return (market or {}).get(side)


def _format_american(am: int | float | None) -> str | None:
    if am is None:
        return None
    am = int(round(am))
    return f"+{am}" if am > 0 else str(am)


def _vig_free(dec_a: float | None, dec_b: float | None) -> tuple[float | None, float | None]:
    """Strip the book's juice off a 2-way pair to recover what the
    market really thinks. Compare model probability to vig-free implied
    probability for a fair edge calc.
    """
    if not dec_a or not dec_b or dec_a <= 1 or dec_b <= 1:
        return (None, None)
    p_a_raw = 1.0 / dec_a
    p_b_raw = 1.0 / dec_b
    total = p_a_raw + p_b_raw
    if total <= 0:
        return (None, None)
    return (p_a_raw / total, p_b_raw / total)


def _inject_status_columns(
    tab: dict,
    bet_type: str,
    gate_passed: set[str] | None,
    gate_notes: dict[str, str] | None,
    edge_thresholds: dict[str, float] | None,
    min_edge_fallback: float,
) -> None:
    """Add `status`, `conviction_pct`, and `status_reason` columns to
    every projection row in a per-market tab, in place.

    `status` is the single human-readable verdict that combines all
    three gates the daily build applies:
      1. Market gate (BRAND_GUIDE: ≥+1% ROI AND Brier <0.246 over
         200+ bets in the rolling backtest)
      2. Market price availability (no synthetic -110)
      3. Per-market edge threshold (e.g. ≥3% for run_line)

    `conviction_pct` surfaces model_prob as a clean number (e.g. 65.4)
    so spreadsheet readers don't have to multiply 0.654 themselves.

    `status_reason` is empty for PLAY rows; for PASS rows it explains
    which gate the row failed.

    Updates tab["projection_columns"] in place so the new fields surface
    in CSV / xlsx output. New columns land near the front of the column
    order (after date/away/home) so they're visible without horizontal
    scrolling.

    Idempotent — safe to call twice on the same tab.
    """
    gate_notes = gate_notes or {}
    edge_thresholds = edge_thresholds or {}

    for row in tab.get("projections", []):
        if gate_passed is not None and bet_type not in gate_passed:
            note = gate_notes.get(bet_type, "failed market gate")
            status, reason = "PASS", f"Market gated off ({note})"
        elif row.get("market_odds_dec") is None:
            status, reason = "PASS", "No market price available"
        elif row.get("edge_pct") is None:
            status, reason = "PASS", "Edge not computed"
        else:
            threshold = edge_thresholds.get(bet_type, min_edge_fallback)
            if row["edge_pct"] >= threshold:
                status, reason = "PLAY", ""
            else:
                status = "PASS"
                reason = (
                    f"Edge {row['edge_pct']:+.2f}% below "
                    f"{threshold:+.2f}% threshold"
                )

        row["status"] = status
        row["status_reason"] = reason
        prob = row.get("model_prob")
        row["conviction_pct"] = (
            round(prob * 100, 1) if prob is not None else None
        )

    cols = list(tab.get("projection_columns", []))
    NEW = ("status", "conviction_pct", "status_reason")
    cols = [c for c in cols if c not in NEW]
    # Insert status + conviction_pct after the first three "where/who"
    # columns (typically date/away/home, or date/team/opponent for
    # team_totals) so they're visible without horizontal scrolling.
    # status_reason goes at the end since it's only useful when you're
    # already looking for "why."
    insert_idx = min(3, len(cols))
    cols = (
        cols[:insert_idx]
        + ["status", "conviction_pct"]
        + cols[insert_idx:]
        + ["status_reason"]
    )
    tab["projection_columns"] = cols


def _build_odds_debug(odds: dict) -> dict:
    """Flat, eyeball-friendly view of the normalized odds payload.

    One entry per game with ML / Run Line / Totals laid out side-by-side
    plus vig-free implied probabilities for sanity checks. Useful when
    Today's Card surfaces an unexpected pick — open this file to see
    exactly what prices the model saw without re-running the build or
    burning another API call.
    """
    games = []
    for g in odds.get("games", []):
        away, home = g["away_team"], g["home_team"]
        record: dict = {
            "game": f"{away} @ {home}",
            "first_pitch": g.get("commence_time"),
        }

        ml = g.get("moneyline") or {}
        away_ml = ml.get("away")
        home_ml = ml.get("home")
        if away_ml or home_ml:
            block: dict = {}
            if away_ml:
                block[away] = {
                    "american": _format_american(away_ml["american"]),
                    "decimal": away_ml["decimal"],
                    "book": away_ml["book"],
                }
            if home_ml:
                block[home] = {
                    "american": _format_american(home_ml["american"]),
                    "decimal": home_ml["decimal"],
                    "book": home_ml["book"],
                }
            if away_ml and home_ml:
                p_a, p_h = _vig_free(away_ml["decimal"], home_ml["decimal"])
                if p_a is not None:
                    block["vig_free"] = {
                        away: f"{p_a * 100:.1f}%",
                        home: f"{p_h * 100:.1f}%",
                    }
            record["moneyline"] = block

        rl_offers = g.get("run_line") or []
        if rl_offers:
            block = {}
            for o in rl_offers:
                team_code = away if o["team"] == "away" else home
                point = o["point"]
                sign = "+" if point > 0 else ""
                key = f"{team_code} {sign}{point}"
                block[key] = {
                    "american": _format_american(o["american"]),
                    "decimal": o["decimal"],
                    "book": o["book"],
                }
            record["run_line"] = block

        total_offers = g.get("totals") or []
        if total_offers:
            offers = []
            for offer in total_offers:
                line = offer.get("point")
                over = offer.get("over")
                under = offer.get("under")
                t: dict = {"line": line}
                if over:
                    t["OVER"] = {
                        "american": _format_american(over["american"]),
                        "decimal": over["decimal"],
                        "book": over["book"],
                    }
                if under:
                    t["UNDER"] = {
                        "american": _format_american(under["american"]),
                        "decimal": under["decimal"],
                        "book": under["book"],
                    }
                if over and under:
                    p_o, p_u = _vig_free(over["decimal"], under["decimal"])
                    if p_o is not None:
                        t["vig_free"] = {
                            "OVER": f"{p_o * 100:.1f}%",
                            "UNDER": f"{p_u * 100:.1f}%",
                        }
                offers.append(t)
            record["totals"] = offers

        games.append(record)

    return {
        "fetched_at": odds.get("fetched_at"),
        "source": odds.get("source"),
        "game_count": len(games),
        "games": games,
    }


def _write_odds_debug_csv(path: Path, debug: dict) -> None:
    """Flat row-per-offer CSV for quick filtering in Excel/Sheets."""
    rows = []
    for g in debug["games"]:
        game = g["game"]
        ml = g.get("moneyline") or {}
        ml_vig = ml.get("vig_free", {})
        for team, info in ml.items():
            if team == "vig_free":
                continue
            rows.append({
                "game": game, "market": "moneyline", "side": team,
                "american": info["american"], "decimal": info["decimal"],
                "book": info["book"], "vig_free_pct": ml_vig.get(team),
            })
        for side, info in (g.get("run_line") or {}).items():
            rows.append({
                "game": game, "market": "run_line", "side": side,
                "american": info["american"], "decimal": info["decimal"],
                "book": info["book"], "vig_free_pct": None,
            })
        for total_offer in (g.get("totals") or []):
            line = total_offer["line"]
            vf = total_offer.get("vig_free", {})
            for side in ("OVER", "UNDER"):
                info = total_offer.get(side)
                if not info:
                    continue
                rows.append({
                    "game": game, "market": "totals",
                    "side": f"{side} {line}",
                    "american": info["american"], "decimal": info["decimal"],
                    "book": info["book"], "vig_free_pct": vf.get(side),
                })

    cols = ["game", "market", "side", "american", "decimal", "book", "vig_free_pct"]
    with path.open("w", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=cols, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)


def _market_gate(
    backtest_summary: list[dict] | None,
    min_bets: int = DEFAULT_MIN_GATE_BETS,
    min_roi: float = DEFAULT_MIN_GATE_ROI,
    max_brier: float = DEFAULT_MAX_GATE_BRIER,
) -> tuple[set[str] | None, dict[str, str]]:
    """Apply the BRAND_GUIDE market-gating rule to backtest summary rows.

    Returns (passed_set, notes_dict). passed_set is None when there's no
    backtest data (cold start — let everything through). notes_dict maps
    bet_type → reason it didn't pass, useful for surfacing in the
    Today's Card section title so users see WHY a market was excluded.
    """
    if not backtest_summary:
        return None, {}

    passed: set[str] = set()
    notes: dict[str, str] = {}
    for row in backtest_summary:
        bt = row.get("bet_type")
        if not bt:
            continue
        bets = row.get("bets", 0) or 0
        roi = row.get("roi_pct") if row.get("roi_pct") is not None else 0.0
        brier = row.get("brier")

        if bets < min_bets:
            notes[bt] = f"sample {bets}<{min_bets}"
            continue
        if roi < min_roi:
            notes[bt] = f"ROI {roi:+.2f}%"
            continue
        if brier is None or brier >= max_brier:
            brier_str = f"{brier:.4f}" if brier is not None else "n/a"
            notes[bt] = f"Brier {brier_str}"
            continue
        passed.add(bt)
    return passed, notes


def _sp_fields(p: dict) -> dict:
    """Pull away/home SP + bullpen names and factors out of a projection."""
    away_sp = p.get("away_sp") or {}
    home_sp = p.get("home_sp") or {}
    away_bp = p.get("away_bp") or {}
    home_bp = p.get("home_bp") or {}
    return {
        "away_sp_name": away_sp.get("name"),
        "home_sp_name": home_sp.get("name"),
        "away_sp_era": away_sp.get("era"),
        "home_sp_era": home_sp.get("era"),
        "away_sp_fip": away_sp.get("fip"),
        "home_sp_fip": home_sp.get("fip"),
        # Phase 2B: blended factor + recent-form context for transparency.
        "away_sp_recent_fip": away_sp.get("recent_fip"),
        "home_sp_recent_fip": home_sp.get("recent_fip"),
        "away_sp_factor": away_sp.get("factor"),
        "home_sp_factor": home_sp.get("factor"),
        "away_bp_era": away_bp.get("era"),
        "home_bp_era": home_bp.get("era"),
        "away_bp_factor": away_bp.get("factor"),
        "home_bp_factor": home_bp.get("factor"),
    }


def _weather_fields(p: dict) -> dict:
    """Pull weather context out of a projection."""
    w = p.get("weather") or {}
    return {
        "venue": w.get("venue"),
        "temp_f": w.get("temp_f"),
        "wind_mph": w.get("wind_mph"),
        "wind_dir": w.get("wind_dir"),
        "weather_factor": w.get("factor"),
    }


def _lineup_fields(p: dict) -> dict:
    """Pull lineup context out of a projection — star count + scratches."""
    a = p.get("away_lineup") or {}
    h = p.get("home_lineup") or {}
    return {
        "away_lineup_posted": a.get("lineup_posted"),
        "home_lineup_posted": h.get("lineup_posted"),
        "away_stars_present": (
            f"{a.get('stars_present', '?')}/{a.get('stars_total', '?')}"
            if a else None
        ),
        "home_stars_present": (
            f"{h.get('stars_present', '?')}/{h.get('stars_total', '?')}"
            if h else None
        ),
        "away_missing_stars": ", ".join(a.get("missing_stars") or []) or None,
        "home_missing_stars": ", ".join(h.get("missing_stars") or []) or None,
        "away_lineup_factor": a.get("factor"),
        "home_lineup_factor": h.get("factor"),
    }


def _rl_market_price(rl_offers: list[dict], side: str, point: float) -> dict | None:
    """Find the run-line price for (side, point); tolerate small numeric drift."""
    for o in rl_offers or []:
        if o["team"] == side and abs(o["point"] - point) < 0.01:
            return o
    return None


def fetch_slate(date: str) -> list[dict]:
    """Fetch every scheduled MLB game for a date (any status)."""
    url = (
        "https://statsapi.mlb.com/api/v1/schedule"
        f"?sportId=1&date={date}"
        "&fields=dates,date,games,gamePk,gameDate,status,detailedState,"
        "teams,away,home,team,id,name"
    )
    resp = requests.get(url, timeout=30)
    resp.raise_for_status()
    data = resp.json()

    slate = []
    for date_obj in data.get("dates", []):
        for game in date_obj.get("games", []):
            try:
                away_id = game["teams"]["away"]["team"]["id"]
                home_id = game["teams"]["home"]["team"]["id"]
            except KeyError:
                continue
            slate.append({
                "date": date_obj["date"],
                "game_pk": game.get("gamePk"),
                "game_time": game.get("gameDate"),
                "status": game.get("status", {}).get("detailedState"),
                "away_team": TEAM_MAP.get(away_id, str(away_id)),
                "home_team": TEAM_MAP.get(home_id, str(home_id)),
            })
    return slate


class DailySpreadsheet:
    """Builds and writes the daily MLB game-results spreadsheet."""

    BET_TABS = (
        "todays_card",
        "moneyline",
        "run_line",
        "totals",
        "first_5",
        "first_inning",
        "team_totals",
        "backtest",
    )

    def __init__(
        self,
        season: int = SEASON_DEFAULT,
        target_date: str | None = None,
        output_dir: Path | None = None,
        odds_api_key: str | None = None,
        skip_odds: bool = False,
        min_edge_pct: float = DEFAULT_MIN_EDGE_PCT,
        top_n: int = DEFAULT_TOP_N,
        edge_thresholds: dict[str, float] | None = None,
        portfolio_cap_per_game: float = DEFAULT_PORTFOLIO_CAP_PER_GAME,
        gate_min_bets: int = DEFAULT_MIN_GATE_BETS,
        gate_min_roi: float = DEFAULT_MIN_GATE_ROI,
        gate_max_brier: float = DEFAULT_MAX_GATE_BRIER,
        skip_market_gate: bool = False,
        include_backfill: bool = False,
        backfill_dir: Path | None = None,
    ):
        self.season = season
        self.target_date = target_date or _today_et()
        self.output_dir = Path(output_dir) if output_dir else DEFAULT_OUTPUT_DIR
        self.scraper = MLBGameScraper()
        self.odds_scraper = MLBOddsScraper(
            api_key=odds_api_key,
            quota_log_path=self.output_dir / "quota_log.json",
        )
        self.pitcher_scraper = MLBPitcherScraper(season=season)
        self.weather_scraper = MLBWeatherScraper()
        self.lineup_scraper = MLBLineupScraper(season=season)
        self.skip_odds = skip_odds
        self.min_edge_pct = min_edge_pct
        self.top_n = top_n
        self.edge_thresholds = edge_thresholds or dict(DEFAULT_EDGE_THRESHOLDS_BY_MARKET)
        self.portfolio_cap_per_game = portfolio_cap_per_game
        self.gate_min_bets = gate_min_bets
        self.gate_min_roi = gate_min_roi
        self.gate_max_brier = gate_max_brier
        self.skip_market_gate = skip_market_gate
        self.include_backfill = include_backfill
        self.backfill_dir = (
            backfill_dir
            if backfill_dir is not None
            else REPO_ROOT / "data" / "backfill" / "mlb"
        )

    def _discover_backfill_seasons(self) -> list[int]:
        """Sorted list of seasons that have a games.json file on disk.
        Skips the current season (live backfill is fetched separately
        and appended to avoid duplicate games)."""
        if not self.backfill_dir.exists():
            return []
        out: list[int] = []
        for child in self.backfill_dir.iterdir():
            if child.is_dir() and child.name.isdigit():
                if (child / "games.json").exists():
                    season = int(child.name)
                    if season != self.season:  # current goes through live fetch
                        out.append(season)
        return sorted(out)

    # --------- data assembly ------------------------------------------------

    def collect(self) -> dict:
        """Pull backfill + slate, compute projections, return structured data."""
        start = SEASON_OPENING_DAY.format(season=self.season)
        end = (
            datetime.strptime(self.target_date, "%Y-%m-%d") - timedelta(days=1)
        ).strftime("%Y-%m-%d")

        print(f"  Fetching backfill {start} -> {end}...")
        backfill = self.scraper.fetch_schedule(start, end)
        print(f"    {len(backfill)} completed games")

        print(f"  Fetching slate for {self.target_date}...")
        slate = fetch_slate(self.target_date)
        print(f"    {len(slate)} scheduled games")

        print("  Fetching probable starting pitchers (with prior-season xwOBA blend)...")
        # The splits_loader exposes prior-season Statcast xwOBA-against
        # to the pitcher scraper as a stabilizing prior. When the loader
        # has the data on disk for a given pitcher with >=100 PA last
        # season, it blends 30% prior-xwOBA factor + 70% current
        # ERA/FIP-based factor; otherwise the current factor is used
        # unchanged. See scrapers/mlb/mlb_pitcher_scraper.py:
        # blend_with_xwoba for the math.
        splits_loader = SplitsLoader(self.backfill_dir)
        try:
            sp_map = self.pitcher_scraper.fetch_factors_for_slate(
                slate, splits_loader=splits_loader,
            )
        except Exception as e:
            print(f"    SP fetch failed ({type(e).__name__}: {e}); proceeding without SP adjustment")
            sp_map = {}
        for g in slate:
            sp = sp_map.get(g.get("game_pk"), {})
            g["away_sp"] = sp.get("away")
            g["home_sp"] = sp.get("home")
        named = sum(
            1 for g in slate
            if (g.get("away_sp") or {}).get("name")
            and (g.get("home_sp") or {}).get("name")
        )
        # Tally xwOBA coverage so the workflow log shows whether the
        # Statcast prior actually contributed today.
        xwoba_hits = sum(
            1 for sides in sp_map.values()
            for side in sides.values()
            if side.get("prior_xwoba") is not None
        )
        sp_total = sum(
            1 for sides in sp_map.values()
            for side in sides.values()
            if side.get("id") is not None
        )
        print(f"    {named}/{len(slate)} games have both probable SPs identified")
        print(f"    {xwoba_hits}/{sp_total} probable SPs got prior-season xwOBA data")
        if xwoba_hits == 0 and sp_total > 0:
            # Surface the why so "0/N" is never a silent failure. The
            # SplitsLoader records the exact files it tried to load and
            # didn't find — print the first one (typically all share the
            # same root cause: prior-season backfill not on disk).
            try:
                report = splits_loader.diagnostic_report()
                missing = report.get("missing_files") or []
                if missing:
                    first = missing[0]
                    print(
                        f"    xwOBA blend inactive — {first['kind']} file "
                        f"missing at {first['path']}. "
                        f"Run scripts/bootstrap_mlb_backfill.sh to populate "
                        f"data/backfill/mlb/<season>/ then re-run."
                    )
                elif not report.get("backfill_dir_exists"):
                    print(
                        f"    xwOBA blend inactive — backfill_dir "
                        f"{report['backfill_dir']} does not exist. "
                        f"Run scripts/bootstrap_mlb_backfill.sh."
                    )
            except Exception:
                pass

        print("  Fetching team bullpen factors (season + last-3-day workload)...")
        team_codes = sorted({g["away_team"] for g in slate} | {g["home_team"] for g in slate})
        try:
            bp_map = self.pitcher_scraper.fetch_bullpen_factors(
                team_codes,
                target_date=self.target_date,
                include_workload=True,
                lookback_days=3,
            )
        except Exception as e:
            print(f"    Bullpen fetch failed ({type(e).__name__}: {e}); proceeding without BP adjustment")
            bp_map = {}
        for g in slate:
            g["away_bp"] = bp_map.get(g["away_team"])
            g["home_bp"] = bp_map.get(g["home_team"])
        bp_seen = sum(
            1 for code, info in bp_map.items()
            if info and info.get("era") is not None
        )
        # Surface the fatigue signal in the workflow log so we can see when
        # the workload feature is actually pulling factors away from
        # season-aggregate. Threshold of 1.05 ≈ at least ~3 IP above
        # normal in the last 3 days.
        tired = sorted(
            (code for code, info in bp_map.items()
             if info and info.get("fatigue_factor", 1.0) >= 1.05),
            key=lambda c: bp_map[c]["fatigue_factor"], reverse=True,
        )
        print(f"    {bp_seen}/{len(team_codes)} teams returned bullpen stats")
        if tired:
            details = ", ".join(
                f"{c}({bp_map[c]['bp_ip_recent']}IP→×{bp_map[c]['fatigue_factor']})"
                for c in tired
            )
            print(f"    Tired bullpens: {details}")

        print("  Fetching weather for each home venue...")
        try:
            weather_map = self.weather_scraper.fetch_for_slate(slate)
        except Exception as e:
            print(f"    Weather fetch failed ({type(e).__name__}: {e}); proceeding with neutral weather")
            weather_map = {}
        for g in slate:
            g["weather"] = weather_map.get(g.get("game_pk"))
        wx_with_temp = sum(
            1 for w in weather_map.values()
            if w and w.get("temp_f") is not None
        )
        domes = sum(1 for w in weather_map.values() if w and w.get("dome"))
        print(f"    {wx_with_temp} outdoor games with weather, {domes} domes")

        print("  Fetching day-of lineups + team star hitters...")
        try:
            lineup_map = self.lineup_scraper.fetch_for_slate(slate)
        except Exception as e:
            print(f"    Lineup fetch failed ({type(e).__name__}: {e}); proceeding with neutral factors")
            lineup_map = {}
        posted = 0
        scratched = 0
        for g in slate:
            entry = lineup_map.get(g.get("game_pk")) or {}
            g["away_lineup"] = entry.get("away")
            g["home_lineup"] = entry.get("home")
            for side_data in (entry.get("away"), entry.get("home")):
                if side_data and side_data.get("lineup_posted"):
                    posted += 1
                if side_data and side_data.get("missing_stars"):
                    scratched += len(side_data["missing_stars"])
        print(f"    {posted}/{len(slate) * 2} lineups posted, {scratched} star scratches detected")

        if self.skip_odds:
            odds = {"fetched_at": None, "source": "skipped", "games": []}
        else:
            print("  Fetching market odds...")
            odds = self.odds_scraper.fetch()
            print(f"    {odds['source']} -> {len(odds['games'])} priced games")

        if self.include_backfill:
            backfill_seasons = self._discover_backfill_seasons()
            print(
                f"  Running backtest with multi-season backfill "
                f"(seasons {backfill_seasons} on disk + current)..."
            )
            backtest_engine = BacktestEngine.from_multi_season(
                backfill_dir=self.backfill_dir,
                seasons=backfill_seasons,
                current_season_games=backfill,
            )
            print(
                f"    Combined sample: {len(backtest_engine.games):,} games "
                f"({getattr(backtest_engine, '_seasons_loaded', {})})"
            )
        else:
            print("  Running backtest (current season only)...")
            backtest_engine = BacktestEngine(backfill)

        backtest = backtest_engine.run()
        cal = backtest["calibration"]
        print(f"    {backtest['overall']['bets']:,} graded bets, "
              f"hit rate {backtest['overall']['hit_rate']}%, "
              f"units {backtest['overall']['units_pl']:+.2f}")
        print(f"    calibrated SDs: total {cal['total_sd']}, margin {cal['margin_sd']}, "
              f"team_total {cal['team_total_sd']}, ML slope {cal['win_prob_slope']}")

        print("  Building projection model with calibrated constants...")
        model = ProjectionModel(backfill, calibration=cal)
        projections = model.project_slate(slate)

        # Optional: hand off NRFI/YRFI probabilities to v1's elite NRFI
        # engine (XGBoost + LightGBM + SHAP + MC) when its runtime is
        # available. Gated by EDGE_FEATURE_NRFI_BRIDGE so the verbatim
        # projector remains the default during the polish phase.
        try:
            from edge_equation.exporters.mlb._nrfi_bridge import apply_overrides
            nrfi_report = apply_overrides(projections, self.target_date)
            if nrfi_report.get("active"):
                mode = nrfi_report.get("mode", "on")
                if mode == "shadow":
                    deltas = nrfi_report.get("shadow_deltas") or []
                    print(
                        f"  NRFI bridge [shadow]: {nrfi_report['applied']} games measured "
                        f"({nrfi_report['skipped']} skipped, "
                        f"engine={nrfi_report.get('engine_label')}). "
                        f"Projections NOT modified."
                    )
                    for d in deltas:
                        print(
                            f"    {d['matchup']}: projector={d['projector_nrfi']:.3f} "
                            f"engine={d['engine_nrfi']:.3f} delta={d['delta']:+.3f}"
                        )
                else:
                    print(
                        f"  NRFI bridge [on]: applied {nrfi_report['applied']} overrides "
                        f"({nrfi_report['skipped']} skipped, "
                        f"engine={nrfi_report.get('engine_label')})"
                    )
            else:
                print(f"  NRFI bridge: inactive ({nrfi_report.get('reason', 'n/a')}); "
                      f"using projector's in-house first-inning math")
        except Exception as e:
            print(f"  NRFI bridge: error ({type(e).__name__}: {e}); "
                  f"using projector's in-house first-inning math")

        tabs = {
            "moneyline": self._build_moneyline(backfill, projections, odds),
            "run_line": self._build_run_line(backfill, projections, odds),
            "totals": self._build_totals(backfill, projections, odds, model.total_sd),
            "first_5": self._build_first_5(backfill, projections, odds),
            "first_inning": self._build_first_inning(backfill, projections, odds),
            "team_totals": self._build_team_totals(
                backfill, projections, odds, model.team_total_sd
            ),
        }
        # Apply BRAND_GUIDE market gate from the backtest summary, then
        # build Today's Card with that filter active.
        gate_passed, gate_notes = (None, {}) if self.skip_market_gate else _market_gate(
            backtest.get("summary_by_bet_type"),
            min_bets=self.gate_min_bets,
            min_roi=self.gate_min_roi,
            max_brier=self.gate_max_brier,
        )
        if gate_passed is not None:
            print(f"  Market gate: {sorted(gate_passed) or '(none passed)'} pass; "
                  f"excluded: {gate_notes}")

        # Inject human-readable status / conviction columns into every
        # per-market tab so CSV + Excel consumers see PLAY/PASS at a
        # glance instead of having to compute the gate decision in
        # their head from edge_pct + market_odds + the gate summary.
        for bet_type in ("moneyline", "run_line", "totals", "first_5",
                         "first_inning", "team_totals"):
            _inject_status_columns(
                tabs[bet_type], bet_type,
                gate_passed=gate_passed,
                gate_notes=gate_notes,
                edge_thresholds=self.edge_thresholds,
                min_edge_fallback=self.min_edge_pct,
            )

        tabs["todays_card"] = self._build_todays_card(
            tabs,
            min_edge_pct=self.min_edge_pct,
            top_n=self.top_n,
            edge_thresholds=self.edge_thresholds,
            portfolio_cap_per_game=self.portfolio_cap_per_game,
            gate_passed=gate_passed,
            gate_notes=gate_notes,
        )

        # Persist today's actionable picks to the CLV log, then load CLV
        # summary stats from any prior closing-snapshot runs.
        tracker = ClvTracker(self.output_dir)
        slate_meta_by_matchup = {
            f"{g['away_team']}@{g['home_team']}": {
                "game_pk": g.get("game_pk"),
                "game_time": g.get("game_time"),
            }
            for g in slate
        }
        added = tracker.record_picks(
            tabs["todays_card"]["projections"],
            odds_source=odds.get("source", "unknown"),
            slate_meta_by_matchup=slate_meta_by_matchup,
        )

        # Grade any unsettled picks whose game has now completed. The
        # backfill we just pulled covers everything through yesterday,
        # so any picks dated yesterday-or-earlier should now have a
        # graded result. Idempotent: already-graded picks are skipped.
        grade_report = tracker.grade_resolved_picks(backfill)
        if grade_report["graded"]:
            print(f"  Graded {grade_report['graded']} resolved pick(s); "
                  f"{grade_report['still_pending']} await game completion")

        clv_summary = tracker.save_summary()
        print(f"  CLV log: +{added} new picks, "
              f"{clv_summary['picks_with_close']}/{clv_summary['picks_total']} "
              f"have closing snapshots, "
              f"{clv_summary['picks_graded']} graded")
        if clv_summary["picks_with_close"]:
            o = clv_summary["overall"]
            print(f"    overall CLV: {o['mean_clv_pct']:+.2f}% mean ({o['positive']}+ / {o['negative']}-)")

        tabs["backtest"] = self._build_backtest(backtest, clv_summary=clv_summary)

        return {
            "generated_at": datetime.utcnow().isoformat() + "Z",
            "season": self.season,
            "today": self.target_date,
            "odds_source": odds["source"],
            "odds_fetched_at": odds["fetched_at"],
            "counts": {
                "backfill_games": len(backfill),
                "slate_games": len(slate),
                "priced_games": len(odds["games"]),
            },
            "tabs": tabs,
            "odds": odds,
            "backtest": backtest,
        }

    # --------- per-tab builders --------------------------------------------

    @staticmethod
    def _build_moneyline(
        backfill: list[dict], projections: list[dict], odds: dict
    ) -> dict:
        proj_rows = []
        for p in projections:
            pick_side = "home" if p["ml_pick"] == p["home_team"] else "away"
            pick_prob = (
                p["home_win_prob"] if pick_side == "home" else p["away_win_prob"]
            )

            game_odds = MLBOddsScraper.find_game(odds, p["away_team"], p["home_team"])
            market = _ml_market_price((game_odds or {}).get("moneyline", {}), pick_side)

            if market:
                adv = kelly_advice(pick_prob, decimal_odds=market["decimal"])
                market_dec = market["decimal"]
                market_am = market["american"]
                book = market["book"]
            else:
                adv = kelly_advice(pick_prob)
                market_dec = None
                market_am = None
                book = None

            proj_rows.append({
                "date": p["date"],
                "away": p["away_team"],
                "home": p["home_team"],
                **_sp_fields(p),
                "away_runs_proj": p["away_runs_proj"],
                "home_runs_proj": p["home_runs_proj"],
                "away_win_prob": p["away_win_prob"],
                "home_win_prob": p["home_win_prob"],
                "ml_pick": p["ml_pick"],
                "model_prob": adv["model_prob"],
                "fair_odds_dec": adv["fair_odds_dec"],
                "market_odds_dec": market_dec,
                "market_odds_american": market_am,
                "book": book,
                "edge_pct": edge_pct(pick_prob, market_dec),
                "kelly_pct": adv["kelly_pct"],
                "kelly_advice": adv["kelly_advice"],
            })
        backfill_rows = [
            {
                "date": g["date"],
                "away": g["away_team"],
                "home": g["home_team"],
                "away_score": g["away_score"],
                "home_score": g["home_score"],
                "ml_winner": g["ml_winner"],
            }
            for g in sorted(backfill, key=lambda g: g["date"], reverse=True)
        ]
        return {
            "title": "Moneyline",
            "projection_columns": [
                "date", "away", "home",
                "away_sp_name", "home_sp_name",
                "away_sp_fip", "home_sp_fip",
                "away_sp_factor", "home_sp_factor",
                "away_bp_factor", "home_bp_factor",
                "away_runs_proj", "home_runs_proj",
                "away_win_prob", "home_win_prob", "ml_pick",
                "model_prob", "fair_odds_dec",
                "market_odds_dec", "market_odds_american", "book",
                "edge_pct", "kelly_pct", "kelly_advice",
            ],
            "backfill_columns": [
                "date", "away", "home", "away_score", "home_score", "ml_winner",
            ],
            "projections": proj_rows,
            "backfill": backfill_rows,
        }

    @staticmethod
    def _build_run_line(
        backfill: list[dict], projections: list[dict], odds: dict
    ) -> dict:
        proj_rows = []
        for p in projections:
            # We bet the projected UNDERDOG at +1.5 (see ProjectionModel
            # comment for the backtest evidence behind this inversion).
            pick_side = "home" if p["rl_pick"] == p["home_team"] else "away"

            game_odds = MLBOddsScraper.find_game(odds, p["away_team"], p["home_team"])
            market = _rl_market_price(
                (game_odds or {}).get("run_line", []), pick_side, 1.5,
            )

            if market:
                adv = kelly_advice(p["rl_cover_prob"], decimal_odds=market["decimal"])
                market_dec = market["decimal"]
                market_am = market["american"]
                book = market["book"]
            else:
                adv = kelly_advice(p["rl_cover_prob"])
                market_dec = None
                market_am = None
                book = None

            proj_rows.append({
                "date": p["date"],
                "away": p["away_team"],
                "home": p["home_team"],
                **_sp_fields(p),
                "rl_fav": p["rl_fav"],
                "rl_pick": p["rl_pick"],          # team we're betting (underdog +1.5)
                "rl_margin_proj": p["rl_margin_proj"],
                "rl_cover_prob": p["rl_cover_prob"],
                "rl_fav_covers_1_5": "YES" if p["rl_fav_covers_1_5"] else "NO",
                "model_prob": adv["model_prob"],
                "fair_odds_dec": adv["fair_odds_dec"],
                "market_odds_dec": market_dec,
                "market_odds_american": market_am,
                "book": book,
                "edge_pct": edge_pct(p["rl_cover_prob"], market_dec),
                "kelly_pct": adv["kelly_pct"],
                "kelly_advice": adv["kelly_advice"],
            })
        backfill_rows = []
        for g in sorted(backfill, key=lambda g: g["date"], reverse=True):
            margin = g["away_score"] - g["home_score"]
            fav = g["away_team"] if margin > 0 else g["home_team"] if margin < 0 else "PUSH"
            backfill_rows.append({
                "date": g["date"],
                "away": g["away_team"],
                "home": g["home_team"],
                "away_score": g["away_score"],
                "home_score": g["home_score"],
                "margin": abs(margin),
                "favorite": fav,
                "fav_covers_1_5": "YES" if g["rl_favorite_covered"] else "NO",
            })
        return {
            "title": "Run Line",
            "projection_columns": [
                "date", "away", "home",
                "away_sp_name", "home_sp_name",
                "away_sp_factor", "home_sp_factor",
                "away_bp_factor", "home_bp_factor",
                "rl_fav", "rl_margin_proj",
                "rl_cover_prob", "rl_fav_covers_1_5",
                "model_prob", "fair_odds_dec",
                "market_odds_dec", "market_odds_american", "book",
                "edge_pct", "kelly_pct", "kelly_advice",
            ],
            "backfill_columns": [
                "date", "away", "home", "away_score", "home_score",
                "margin", "favorite", "fav_covers_1_5",
            ],
            "projections": proj_rows,
            "backfill": backfill_rows,
        }

    @staticmethod
    def _build_totals(
        backfill: list[dict],
        projections: list[dict],
        odds: dict,
        total_sd: float = TOTAL_SD,
    ) -> dict:
        proj_rows = []
        for p in projections:
            row = {
                "date": p["date"],
                "away": p["away_team"],
                "home": p["home_team"],
                **_sp_fields(p),
                **_weather_fields(p),
                **_lineup_fields(p),
                "away_runs_proj": p["away_runs_proj"],
                "home_runs_proj": p["home_runs_proj"],
                "total_proj": p["total_proj"],
            }
            for line in TOTAL_LINES:
                row[f"pick_{line}"] = "OVER" if p["total_proj"] > line else "UNDER"

            game_odds = MLBOddsScraper.find_game(odds, p["away_team"], p["home_team"])
            market_totals = (game_odds or {}).get("totals") or []
            best = _market_total_kelly(p["total_proj"], total_sd, market_totals)
            if best is None:
                best = _best_total_kelly(p["total_proj"], TOTAL_LINES, total_sd)
                row.update({
                    "kelly_line": f"{best['side']} {best['line']}",
                    "model_prob": best["model_prob"],
                    "fair_odds_dec": best["fair_odds_dec"],
                    "market_odds_dec": None,
                    "market_odds_american": None,
                    "book": None,
                    "edge_pct": None,
                    "kelly_pct": best["kelly_pct"],
                    "kelly_advice": best["kelly_advice"],
                })
            else:
                row.update({
                    "kelly_line": f"{best['side']} {best['line']}",
                    "model_prob": best["model_prob"],
                    "fair_odds_dec": best["fair_odds_dec"],
                    "market_odds_dec": best["market_decimal"],
                    "market_odds_american": best["market_american"],
                    "book": best["book"],
                    "edge_pct": best["edge_pct"],
                    "kelly_pct": best["kelly_pct"],
                    "kelly_advice": best["kelly_advice"],
                })
            proj_rows.append(row)

        backfill_rows = []
        for g in sorted(backfill, key=lambda g: g["date"], reverse=True):
            row = {
                "date": g["date"],
                "away": g["away_team"],
                "home": g["home_team"],
                "total_runs": g["total"],
            }
            for line in TOTAL_LINES:
                row[f"result_{line}"] = _ou_label(g["total"], line)
            backfill_rows.append(row)

        return {
            "title": "Totals",
            "projection_columns": [
                "date", "away", "home",
                "away_sp_name", "home_sp_name",
                "away_sp_factor", "home_sp_factor",
                "away_bp_factor", "home_bp_factor",
                "venue", "temp_f", "wind_mph", "wind_dir", "weather_factor",
                "away_stars_present", "home_stars_present",
                "away_missing_stars", "home_missing_stars",
                "away_lineup_factor", "home_lineup_factor",
                "away_runs_proj", "home_runs_proj", "total_proj",
            ] + [f"pick_{l}" for l in TOTAL_LINES] + [
                "kelly_line", "model_prob", "fair_odds_dec",
                "market_odds_dec", "market_odds_american", "book",
                "edge_pct", "kelly_pct", "kelly_advice",
            ],
            "backfill_columns": [
                "date", "away", "home", "total_runs",
            ] + [f"result_{l}" for l in TOTAL_LINES],
            "projections": proj_rows,
            "backfill": backfill_rows,
        }

    @staticmethod
    def _build_first_5(
        backfill: list[dict], projections: list[dict], odds: dict
    ) -> dict:
        proj_rows = []
        for p in projections:
            adv = kelly_advice(p["f5_win_prob"]) if p["f5_pick"] != "PUSH" \
                else {"model_prob": 0.5, "fair_odds_dec": 2.0,
                      "kelly_pct": 0.0, "kelly_advice": "PASS"}
            proj_rows.append({
                "date": p["date"],
                "away": p["away_team"],
                "home": p["home_team"],
                **_sp_fields(p),
                "f5_away_proj": p["f5_away_proj"],
                "f5_home_proj": p["f5_home_proj"],
                "f5_total_proj": p["f5_total_proj"],
                "f5_pick": p["f5_pick"],
                "f5_win_prob": p["f5_win_prob"],
                "model_prob": adv["model_prob"],
                "fair_odds_dec": adv["fair_odds_dec"],
                "kelly_pct": adv["kelly_pct"],
                "kelly_advice": adv["kelly_advice"],
            })
        backfill_rows = [
            {
                "date": g["date"],
                "away": g["away_team"],
                "home": g["home_team"],
                "f5_away": g["f5_away"],
                "f5_home": g["f5_home"],
                "f5_total": g["f5_away"] + g["f5_home"],
                "f5_winner": g["f5_winner"],
            }
            for g in sorted(backfill, key=lambda g: g["date"], reverse=True)
        ]
        return {
            "title": "First 5",
            "projection_columns": [
                "date", "away", "home",
                "away_sp_name", "home_sp_name",
                "away_sp_factor", "home_sp_factor",
                "f5_away_proj", "f5_home_proj",
                "f5_total_proj", "f5_pick", "f5_win_prob",
                "model_prob", "fair_odds_dec", "kelly_pct", "kelly_advice",
            ],
            "backfill_columns": [
                "date", "away", "home", "f5_away", "f5_home", "f5_total", "f5_winner",
            ],
            "projections": proj_rows,
            "backfill": backfill_rows,
        }

    @staticmethod
    def _build_first_inning(
        backfill: list[dict], projections: list[dict], odds: dict
    ) -> dict:
        proj_rows = []
        for p in projections:
            pick_prob = p["nrfi_prob"] if p["nrfi_pick"] == "NRFI" else p["yrfi_prob"]
            adv = kelly_advice(pick_prob)
            proj_rows.append({
                "date": p["date"],
                "away": p["away_team"],
                "home": p["home_team"],
                **_sp_fields(p),
                "f1_total_proj": p["f1_total_proj"],
                "nrfi_prob": p["nrfi_prob"],
                "yrfi_prob": p["yrfi_prob"],
                "nrfi_pick": p["nrfi_pick"],
                "model_prob": adv["model_prob"],
                "fair_odds_dec": adv["fair_odds_dec"],
                "kelly_pct": adv["kelly_pct"],
                "kelly_advice": adv["kelly_advice"],
            })
        backfill_rows = [
            {
                "date": g["date"],
                "away": g["away_team"],
                "home": g["home_team"],
                "f1_away": g["f1_away"],
                "f1_home": g["f1_home"],
                "f1_total": g["f1_away"] + g["f1_home"],
                "nrfi_yrfi": "NRFI" if g["nrfi"] else "YRFI",
            }
            for g in sorted(backfill, key=lambda g: g["date"], reverse=True)
        ]
        return {
            "title": "First Inning",
            "projection_columns": [
                "date", "away", "home",
                "away_sp_name", "home_sp_name",
                "f1_total_proj",
                "nrfi_prob", "yrfi_prob", "nrfi_pick",
                "model_prob", "fair_odds_dec", "kelly_pct", "kelly_advice",
            ],
            "backfill_columns": [
                "date", "away", "home", "f1_away", "f1_home", "f1_total", "nrfi_yrfi",
            ],
            "projections": proj_rows,
            "backfill": backfill_rows,
        }

    @staticmethod
    def _build_team_totals(
        backfill: list[dict],
        projections: list[dict],
        odds: dict,
        team_total_sd: float = TEAM_TOTAL_SD,
    ) -> dict:
        proj_rows = []
        for p in projections:
            for side, team, opp, runs, opp_sp, opp_bp in (
                ("AWAY", p["away_team"], p["home_team"], p["away_runs_proj"],
                 p.get("home_sp"), p.get("home_bp")),
                ("HOME", p["home_team"], p["away_team"], p["home_runs_proj"],
                 p.get("away_sp"), p.get("away_bp")),
            ):
                row = {
                    "date": p["date"],
                    "team": team,
                    "opponent": opp,
                    "side": side,
                    "opp_sp_name": (opp_sp or {}).get("name"),
                    "opp_sp_factor": (opp_sp or {}).get("factor"),
                    "opp_bp_factor": (opp_bp or {}).get("factor"),
                    "team_total_proj": runs,
                }
                for line in TEAM_TOTAL_LINES:
                    row[f"pick_{line}"] = "OVER" if runs > line else "UNDER"
                best = _best_total_kelly(runs, TEAM_TOTAL_LINES, team_total_sd)
                row.update({
                    "kelly_line": f"{best['side']} {best['line']}",
                    "model_prob": best["model_prob"],
                    "fair_odds_dec": best["fair_odds_dec"],
                    "kelly_pct": best["kelly_pct"],
                    "kelly_advice": best["kelly_advice"],
                })
                proj_rows.append(row)

        backfill_rows = []
        for g in sorted(backfill, key=lambda g: g["date"], reverse=True):
            for side, team, opp, runs in (
                ("AWAY", g["away_team"], g["home_team"], g["away_score"]),
                ("HOME", g["home_team"], g["away_team"], g["home_score"]),
            ):
                row = {
                    "date": g["date"],
                    "team": team,
                    "opponent": opp,
                    "side": side,
                    "runs": runs,
                }
                for line in TEAM_TOTAL_LINES:
                    row[f"result_{line}"] = _ou_label(runs, line)
                backfill_rows.append(row)

        return {
            "title": "Team Totals",
            "projection_columns": [
                "date", "team", "opponent", "side",
                "opp_sp_name", "opp_sp_factor", "opp_bp_factor",
                "team_total_proj",
            ] + [f"pick_{l}" for l in TEAM_TOTAL_LINES] + [
                "kelly_line", "model_prob", "fair_odds_dec", "kelly_pct", "kelly_advice",
            ],
            "backfill_columns": [
                "date", "team", "opponent", "side", "runs",
            ] + [f"result_{l}" for l in TEAM_TOTAL_LINES],
            "projections": proj_rows,
            "backfill": backfill_rows,
        }

    @staticmethod
    def _build_todays_card(
        tabs: dict,
        min_edge_pct: float = DEFAULT_MIN_EDGE_PCT,
        top_n: int = DEFAULT_TOP_N,
        edge_thresholds: dict[str, float] | None = None,
        portfolio_cap_per_game: float = DEFAULT_PORTFOLIO_CAP_PER_GAME,
        gate_passed: set[str] | None = None,
        gate_notes: dict[str, str] | None = None,
    ) -> dict:
        """Cross-tab roll-up of the highest-edge plays.

        Inclusion rules:
          1. **Market gate** (BRAND_GUIDE Product Rules): bet_type must
             have passed the rolling backtest gate (≥+1% ROI AND
             Brier <0.246 over 200+ bets). Markets that fail still appear
             in their own tabs; just not on the headline card.
          2. Pick must have a real market price (no synthetic -110).
          3. edge_pct ≥ per-market threshold (or min_edge_pct as fallback).
          4. Sort by edge desc, take top N.
          5. Cap total Kelly% per game at portfolio_cap_per_game (same-game
             bets are correlated; full-Kelly on each over-stakes).

        Picks failing rule 1 are dropped entirely (not visible on card or
        FADE list — the market is "off"). Picks failing rule 3 land in
        the FADE / SKIP list. Picks scaled by rule 5 carry a
        `portfolio_scaled_from` field showing the original.

        gate_passed: set of bet_types that passed the market gate, or
        None when no gating is in effect (cold start / disabled).
        """
        thresholds = edge_thresholds or {}
        gate_notes = gate_notes or {}
        sources = (
            ("moneyline",   "ml_pick"),
            ("run_line",    "rl_pick"),     # underdog +1.5 (post-inversion)
            ("totals",      "kelly_line"),
            ("first_5",     "f5_pick"),
            ("first_inning","nrfi_pick"),
            ("team_totals", "kelly_line"),
        )
        rows = []
        for tab_key, pick_key in sources:
            # Market gate: drop the entire bet_type's contributions if it
            # didn't pass the rolling backtest filter. None means no gate.
            if gate_passed is not None and tab_key not in gate_passed:
                continue
            for r in tabs[tab_key]["projections"]:
                if r.get("kelly_pct") is None:
                    continue
                if "team" in r and "opponent" in r:
                    matchup = f"{r['team']} vs {r['opponent']}"
                else:
                    matchup = f"{r.get('away')}@{r.get('home')}"
                # Pitcher context — game-level tabs have away/home_sp_name,
                # team_totals has opp_sp_name. Surface whichever is present.
                sp_context_parts = []
                if r.get("away_sp_name"):
                    sp_context_parts.append(f"A: {r['away_sp_name']}")
                if r.get("home_sp_name"):
                    sp_context_parts.append(f"H: {r['home_sp_name']}")
                if r.get("opp_sp_name") and not sp_context_parts:
                    sp_context_parts.append(f"vs {r['opp_sp_name']}")

                rows.append({
                    "date": r.get("date"),
                    "matchup": matchup,
                    "starting_pitchers": " · ".join(sp_context_parts) or None,
                    "bet_type": tab_key,
                    "pick": r.get(pick_key),
                    "model_prob": r.get("model_prob"),
                    "fair_odds_dec": r.get("fair_odds_dec"),
                    "market_odds_dec": r.get("market_odds_dec"),
                    "market_odds_american": r.get("market_odds_american"),
                    "book": r.get("book"),
                    "edge_pct": r.get("edge_pct"),
                    "kelly_pct": r.get("kelly_pct"),
                    "kelly_advice": r.get("kelly_advice"),
                })

        priced = [r for r in rows if r.get("edge_pct") is not None]
        priced.sort(key=lambda r: r["edge_pct"], reverse=True)

        def _meets_threshold(r: dict) -> bool:
            t = thresholds.get(r["bet_type"], min_edge_pct)
            return r["edge_pct"] >= t

        plays = [r for r in priced if _meets_threshold(r)]
        if top_n and len(plays) > top_n:
            plays = plays[:top_n]

        # Portfolio cap — same-game bets are correlated; cap the sum of
        # half-Kelly across one matchup at portfolio_cap_per_game (% of
        # bankroll). Scale down proportionally when exceeded and re-tier.
        from collections import defaultdict
        by_game: dict[str, list[dict]] = defaultdict(list)
        for p in plays:
            by_game[p["matchup"]].append(p)
        for matchup, game_plays in by_game.items():
            total_pct = sum(p.get("kelly_pct") or 0 for p in game_plays)
            if total_pct > portfolio_cap_per_game and total_pct > 0:
                scale = portfolio_cap_per_game / total_pct
                for p in game_plays:
                    original = p["kelly_pct"]
                    scaled = round(original * scale, 2)
                    p["portfolio_scaled_from"] = original
                    p["portfolio_total_pct"] = round(total_pct, 2)
                    p["kelly_pct"] = scaled
                    p["kelly_advice"] = tier_from_pct(scaled)

        leans = [
            r for r in priced
            if r["edge_pct"] > 0 and not _meets_threshold(r)
        ]
        no_market = [r for r in rows if r.get("edge_pct") is None]
        negative_edge = [r for r in priced if r["edge_pct"] <= 0]

        scaled_count = sum(
            1 for p in plays if p.get("portfolio_scaled_from") is not None
        )
        threshold_note = "per-market" if thresholds else f"≥ {min_edge_pct:.1f}%"
        scaled_note = (
            f" · {scaled_count} portfolio-scaled" if scaled_count else ""
        )
        if gate_passed is not None:
            gate_in = ", ".join(sorted(gate_passed)) or "none"
            gate_out = ", ".join(
                f"{m}({reason})" for m, reason in sorted(gate_notes.items())
            ) or "none"
            gate_note = f" · markets IN: {gate_in} · OUT: {gate_out}"
        else:
            gate_note = " · gate disabled"

        return {
            "title": "Today's Card",
            "projection_section_title":
                f"PLAYS — edge {threshold_note} (top {top_n}) — "
                f"{len(plays)} of {len(priced)} priced picks{scaled_note}",
            "backfill_section_title":
                f"FADE / SKIP — leans below threshold ({len(leans)}) · "
                f"negative-EV ({len(negative_edge)}) · "
                f"no-market ({len(no_market)}){gate_note}",
            "projection_columns": [
                "date", "matchup", "starting_pitchers",
                "bet_type", "pick", "model_prob",
                "fair_odds_dec", "market_odds_dec", "market_odds_american",
                "book", "edge_pct", "kelly_pct", "kelly_advice",
                "portfolio_scaled_from",
            ],
            "backfill_columns": [
                "date", "matchup", "starting_pitchers",
                "bet_type", "pick", "model_prob",
                "market_odds_dec", "edge_pct", "kelly_pct",
            ],
            "projections": plays,
            "backfill": leans + negative_edge + no_market,
        }

    @staticmethod
    def _build_backtest(backtest: dict, clv_summary: dict | None = None) -> dict:
        overall = backtest["overall"]
        summary_rows = [{
            "scope": "OVERALL",
            "bet_type": "all",
            "bets": overall["bets"],
            "wins": overall["wins"],
            "losses": overall["losses"],
            "pushes": overall["pushes"],
            "hit_rate": overall["hit_rate"],
            "units_pl": overall["units_pl"],
            "roi_pct": overall["roi_pct"],
            "brier": overall.get("brier"),
        }]
        for row in backtest["summary_by_bet_type"]:
            summary_rows.append({"scope": "BY TYPE", **row})

        # CLV rows go below the backtest summary in the same section so
        # users can compare modeled performance vs market-validated edge.
        if clv_summary and clv_summary.get("picks_with_close"):
            o = clv_summary["overall"]
            summary_rows.append({
                "scope": "CLV",
                "bet_type": "all",
                "bets": o["n"],
                "wins": o["positive"],
                "losses": o["negative"],
                "pushes": o["neutral"],
                "hit_rate": round(o["positive"] / o["n"] * 100, 1) if o["n"] else 0.0,
                "units_pl": o["mean_clv_pct"],
                "roi_pct": o["median_clv_pct"],
            })
            for bt, stats in clv_summary["by_bet_type"].items():
                summary_rows.append({
                    "scope": "CLV BY TYPE",
                    "bet_type": bt,
                    "bets": stats["n"],
                    "wins": stats["positive"],
                    "losses": stats["negative"],
                    "pushes": stats["neutral"],
                    "hit_rate":
                        round(stats["positive"] / stats["n"] * 100, 1) if stats["n"] else 0.0,
                    "units_pl": stats["mean_clv_pct"],
                    "roi_pct": stats["median_clv_pct"],
                })

        clv_note = ""
        if clv_summary:
            clv_note = (
                f" · CLV tracked: {clv_summary['picks_with_close']} of "
                f"{clv_summary['picks_total']} logged picks"
            )

        return {
            "title": "Backtest",
            "projection_section_title":
                f"BACKTEST SUMMARY — "
                f"{backtest['first_date']} → {backtest['last_date']}{clv_note}",
            "backfill_section_title": "DAILY P&L (cumulative units, flat 1u @ -110)",
            "projection_columns": [
                "scope", "bet_type", "bets", "wins", "losses", "pushes",
                "hit_rate", "units_pl", "roi_pct", "brier",
            ],
            "backfill_columns": ["date", "daily_units", "cumulative_units"],
            "projections": summary_rows,
            "backfill": list(reversed(backtest["daily_pl"])),
        }

    # --------- writers -----------------------------------------------------

    def write_all(self, data: dict) -> list[Path]:
        """Write JSON, per-tab CSVs, and the multi-tab XLSX."""
        self.output_dir.mkdir(parents=True, exist_ok=True)
        written: list[Path] = []

        json_path = self.output_dir / "mlb_daily.json"
        json_path.write_text(json.dumps(data, indent=2, default=str))
        written.append(json_path)

        odds_payload = data.get("odds")
        if odds_payload is not None:
            odds_path = self.output_dir / "lines.json"
            odds_path.write_text(json.dumps(odds_payload, indent=2, default=str))
            written.append(odds_path)

            debug_view = _build_odds_debug(odds_payload)
            debug_json = self.output_dir / "odds_debug.json"
            debug_json.write_text(json.dumps(debug_view, indent=2, default=str))
            written.append(debug_json)

            debug_csv = self.output_dir / "odds_debug.csv"
            _write_odds_debug_csv(debug_csv, debug_view)
            written.append(debug_csv)

        backtest_payload = data.get("backtest")
        if backtest_payload is not None:
            backtest_path = self.output_dir / "backtest.json"
            backtest_path.write_text(
                json.dumps(backtest_payload, indent=2, default=str)
            )
            written.append(backtest_path)

            cal = backtest_payload.get("calibration")
            if cal:
                cal_path = self.output_dir / "calibration.json"
                cal_path.write_text(json.dumps(cal, indent=2, default=str))
                written.append(cal_path)

        for key in self.BET_TABS:
            tab = data["tabs"][key]
            csv_path = self.output_dir / f"{key}.csv"
            self._write_tab_csv(csv_path, tab)
            written.append(csv_path)

        xlsx_path = self.output_dir / "mlb_daily.xlsx"
        self._write_xlsx(xlsx_path, data)
        written.append(xlsx_path)

        return written

    @staticmethod
    def _write_tab_csv(path: Path, tab: dict) -> None:
        all_cols = ["section"] + sorted(
            set(tab["projection_columns"]) | set(tab["backfill_columns"]),
            key=lambda c: (
                tab["projection_columns"].index(c)
                if c in tab["projection_columns"]
                else len(tab["projection_columns"]) + tab["backfill_columns"].index(c)
            ),
        )
        with path.open("w", newline="") as f:
            writer = csv.DictWriter(f, fieldnames=all_cols, extrasaction="ignore")
            writer.writeheader()
            for row in tab["projections"]:
                writer.writerow({"section": "projection", **row})
            for row in tab["backfill"]:
                writer.writerow({"section": "backfill", **row})

    @staticmethod
    def _grade_formula(edge_col: str, row_idx: int) -> str:
        """In-cell grade ladder D→A+ keyed off the row's edge_pct cell.
        Mirrors src/edge_equation/math/scoring.py thresholds restated in
        percent (because the workbook stores edge as a percent value
        like 4.5, not a fraction 0.045)."""
        e = f"{edge_col}{row_idx}"
        return (
            f'=IF({e}="","",'
            f'IF({e}>=8,"A+",'
            f'IF({e}>=5,"A",'
            f'IF({e}>=3,"B",'
            f'IF({e}>=0,"C",'
            f'IF({e}>=-3,"D","F"))))))'
        )

    @staticmethod
    def _kelly_advice_formula(kelly_col: str, row_idx: int) -> str:
        """In-cell mirror of exporters.mlb.kelly._tier(). Resolves the
        same PASS / 0.5u / 1u / 2u / 3u tier as the Python helper but
        recomputes if the kelly_pct cell is edited by hand."""
        k = f"{kelly_col}{row_idx}"
        return (
            f'=IF({k}="","PASS",'
            f'IF({k}<=0.5,"PASS",'
            f'IF({k}<=1.5,"0.5u",'
            f'IF({k}<=3,"1u",'
            f'IF({k}<=5,"2u","3u")))))'
        )

    def _write_xlsx(self, path: Path, data: dict) -> None:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter

        wb = Workbook()
        wb.remove(wb.active)

        # Brand metadata — visible in File > Info / Properties without
        # disturbing cell layout. The "Facts. Not Feelings." tagline is
        # the brand-mandatory string from compliance.brand.BRAND_TAGLINE.
        try:
            from edge_equation.compliance.brand import BRAND_FULL, BRAND_TAGLINE
            wb.properties.title = f"Edge Equation — MLB Daily ({data.get('today', '')})"
            wb.properties.subject = BRAND_TAGLINE
            wb.properties.creator = BRAND_FULL
            wb.properties.keywords = "MLB, picks, NegBin, Kelly, edge"
        except Exception:
            # Brand stamp must never block the build.
            pass

        title_font = Font(bold=True, size=14, color="FFFFFF")
        title_fill = PatternFill("solid", fgColor="1F4E78")
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill("solid", fgColor="2E75B6")
        backfill_title_fill = PatternFill("solid", fgColor="2E5984")

        for key in self.BET_TABS:
            tab = data["tabs"][key]
            ws = wb.create_sheet(title=tab["title"])

            proj_title = tab.get(
                "projection_section_title", f"PROJECTIONS — {data['today']}"
            )
            backfill_title = tab.get(
                "backfill_section_title",
                f"BACKFILL — {data['season']} season-to-date "
                f"({data['counts']['backfill_games']} games)",
            )

            # Append a synthetic "grade" column to the xlsx view when the
            # tab has edge_pct (so the formula has something to reference).
            # Only the xlsx is affected — CSV / JSON outputs use the
            # original projection_columns list.
            base_cols = list(tab["projection_columns"])
            xlsx_cols = list(base_cols)
            if "edge_pct" in xlsx_cols and "grade" not in xlsx_cols:
                xlsx_cols.append("grade")
            col_letter = {
                name: get_column_letter(i + 1) for i, name in enumerate(xlsx_cols)
            }

            ws.cell(row=1, column=1, value=proj_title)
            ws.cell(row=1, column=1).font = title_font
            ws.cell(row=1, column=1).fill = title_fill
            ws.merge_cells(
                start_row=1, start_column=1,
                end_row=1, end_column=max(len(xlsx_cols), 1),
            )

            for c, col in enumerate(xlsx_cols, 1):
                cell = ws.cell(row=2, column=c, value=col)
                cell.font = header_font
                cell.fill = header_fill
            for r, row in enumerate(tab["projections"], 3):
                for c, col in enumerate(xlsx_cols, 1):
                    if col == "grade" and "edge_pct" in col_letter:
                        ws.cell(row=r, column=c,
                                value=self._grade_formula(col_letter["edge_pct"], r))
                    elif col == "kelly_advice" and "kelly_pct" in col_letter:
                        ws.cell(row=r, column=c,
                                value=self._kelly_advice_formula(col_letter["kelly_pct"], r))
                    else:
                        ws.cell(row=r, column=c, value=row.get(col))

            backfill_start = 3 + len(tab["projections"]) + 1
            ws.cell(row=backfill_start, column=1, value=backfill_title)
            ws.cell(row=backfill_start, column=1).font = title_font
            ws.cell(row=backfill_start, column=1).fill = backfill_title_fill
            ws.merge_cells(
                start_row=backfill_start, start_column=1,
                end_row=backfill_start, end_column=max(len(tab["backfill_columns"]), 1),
            )

            for c, col in enumerate(tab["backfill_columns"], 1):
                cell = ws.cell(row=backfill_start + 1, column=c, value=col)
                cell.font = header_font
                cell.fill = header_fill
            for r, row in enumerate(tab["backfill"], backfill_start + 2):
                for c, col in enumerate(tab["backfill_columns"], 1):
                    ws.cell(row=r, column=c, value=row.get(col))

            for col_idx, col in enumerate(
                set(xlsx_cols) | set(tab["backfill_columns"]), 1
            ):
                ws.column_dimensions[
                    ws.cell(row=2, column=col_idx).column_letter
                ].width = max(12, len(col) + 2)

            ws.freeze_panes = "A3"

        wb.save(path)

    # --------- git ---------------------------------------------------------

    def commit_and_push(
        self,
        files: Iterable[Path],
        branch: str | None = None,
    ) -> bool:
        """Stage, commit, and push the written files. Returns True on success."""
        rel = [str(p.relative_to(REPO_ROOT)) for p in files]

        try:
            subprocess.run(
                ["git", "-C", str(REPO_ROOT), "add", *rel],
                check=True, capture_output=True, text=True,
            )
            status = subprocess.run(
                ["git", "-C", str(REPO_ROOT), "status", "--porcelain", *rel],
                check=True, capture_output=True, text=True,
            )
            if not status.stdout.strip():
                print("  No changes to commit.")
                return True

            msg = f"Daily MLB spreadsheet — {self.target_date}"
            subprocess.run(
                ["git", "-C", str(REPO_ROOT), "commit", "-m", msg],
                check=True, capture_output=True, text=True,
            )
            print(f"  Committed: {msg}")

            push_cmd = ["git", "-C", str(REPO_ROOT), "push"]
            if branch:
                push_cmd += ["-u", "origin", branch]
            subprocess.run(push_cmd, check=True, capture_output=True, text=True)
            print("  Pushed.")
            return True
        except subprocess.CalledProcessError as e:
            print(f"  git failed: {e.stderr or e.stdout}")
            return False


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description="MLB Daily Spreadsheet Exporter")
    parser.add_argument("--season", type=int, default=SEASON_DEFAULT)
    parser.add_argument(
        "--date", type=str, default=None,
        help="Target date in YYYY-MM-DD (defaults to today, ET)",
    )
    parser.add_argument(
        "--output-dir", type=str, default=str(DEFAULT_OUTPUT_DIR),
        help="Directory to write outputs into",
    )
    parser.add_argument(
        "--push", action="store_true", default=False,
        help="After writing, git add/commit/push the new files",
    )
    parser.add_argument(
        "--branch", type=str, default=None,
        help="Branch to push to (only with --push)",
    )
    parser.add_argument(
        "--odds-api-key", type=str, default=None,
        help="The Odds API key (overrides ODDS_API_KEY env var)",
    )
    parser.add_argument(
        "--no-odds", action="store_true", default=False,
        help="Skip the odds fetch entirely; use -110 default for all Kelly sizing",
    )
    parser.add_argument(
        "--min-edge", type=float, default=DEFAULT_MIN_EDGE_PCT,
        help=f"Minimum edge_pct (model − implied market) to include in "
             f"Today's Card (default {DEFAULT_MIN_EDGE_PCT})",
    )
    parser.add_argument(
        "--top-n", type=int, default=DEFAULT_TOP_N,
        help=f"Cap Today's Card at this many plays (default {DEFAULT_TOP_N}; "
             f"set 0 to disable)",
    )
    parser.add_argument(
        "--portfolio-cap", type=float, default=DEFAULT_PORTFOLIO_CAP_PER_GAME,
        help=f"Max sum of half-Kelly%% across all bets on one game "
             f"(default {DEFAULT_PORTFOLIO_CAP_PER_GAME})",
    )
    parser.add_argument(
        "--edge-threshold", action="append", default=[],
        metavar="MARKET=PCT",
        help="Override per-market edge threshold, e.g. --edge-threshold totals=2.0 "
             "--edge-threshold moneyline=5.0. Repeat for multiple markets.",
    )
    parser.add_argument(
        "--gate-min-bets", type=int, default=DEFAULT_MIN_GATE_BETS,
        help=f"Minimum backtest sample size for a market to clear the "
             f"BRAND_GUIDE gate (default {DEFAULT_MIN_GATE_BETS}).",
    )
    parser.add_argument(
        "--gate-min-roi", type=float, default=DEFAULT_MIN_GATE_ROI,
        help=f"Minimum backtest ROI%% for a market to clear the "
             f"BRAND_GUIDE gate (default {DEFAULT_MIN_GATE_ROI}).",
    )
    parser.add_argument(
        "--gate-max-brier", type=float, default=DEFAULT_MAX_GATE_BRIER,
        help=f"Maximum backtest Brier for a market to clear the "
             f"BRAND_GUIDE gate (default {DEFAULT_MAX_GATE_BRIER}).",
    )
    parser.add_argument(
        "--no-market-gate", action="store_true", default=False,
        help="Disable the market gate entirely (useful for cold-start runs "
             "or A/B comparisons; defeats the BRAND_GUIDE rule).",
    )
    parser.add_argument(
        "--include-backfill", action="store_true", default=False,
        help="Include multi-season historical games from data/backfill/mlb/ "
             "in the backtest. Larger sample → tighter calibration + more "
             "statistical power for the gate. Adds ~1-2 minutes to the run.",
    )
    parser.add_argument(
        "--backfill-dir", type=str, default=None,
        help="Override the backfill directory (default: <repo>/data/backfill/mlb).",
    )
    args = parser.parse_args(argv)

    # Feature-flag gate runs AFTER argparse so --help / --version still work
    # for operators inspecting the CLI surface. The gate only blocks actual
    # pipeline execution (network calls, file writes, git push).
    if not _feature_flag_on():
        print(
            f"[mlb-daily] feature flag {FEATURE_FLAG_ENV}=off — skipping. "
            f"Set {FEATURE_FLAG_ENV}=on to run.",
            file=sys.stderr,
        )
        return 0

    # Parse --edge-threshold overrides into a dict, falling back to defaults.
    thresholds = dict(DEFAULT_EDGE_THRESHOLDS_BY_MARKET)
    for spec in args.edge_threshold:
        try:
            market, pct = spec.split("=", 1)
            thresholds[market.strip()] = float(pct)
        except (ValueError, TypeError):
            print(f"  Ignoring malformed --edge-threshold: {spec!r}")

    spreadsheet = DailySpreadsheet(
        season=args.season,
        target_date=args.date,
        output_dir=Path(args.output_dir),
        odds_api_key=args.odds_api_key,
        skip_odds=args.no_odds,
        min_edge_pct=args.min_edge,
        top_n=args.top_n,
        edge_thresholds=thresholds,
        portfolio_cap_per_game=args.portfolio_cap,
        gate_min_bets=args.gate_min_bets,
        gate_min_roi=args.gate_min_roi,
        gate_max_brier=args.gate_max_brier,
        skip_market_gate=args.no_market_gate,
        include_backfill=args.include_backfill,
        backfill_dir=Path(args.backfill_dir) if args.backfill_dir else None,
    )

    print(f"MLB Daily Spreadsheet — target {spreadsheet.target_date}")
    data = spreadsheet.collect()
    written = spreadsheet.write_all(data)

    print(f"Wrote {len(written)} files to {spreadsheet.output_dir}:")
    for p in written:
        print(f"  - {p.relative_to(REPO_ROOT)}")

    if args.push:
        ok = spreadsheet.commit_and_push(written, branch=args.branch)
        return 0 if ok else 1
    return 0


if __name__ == "__main__":
    sys.exit(main())
