"""
End-to-end test: orchestrator's _write_xlsx round-trip.

Builds a synthetic `data` dict matching the shape collect() produces,
calls DailySpreadsheet._write_xlsx, reads the resulting xlsx back with
openpyxl, and asserts the in-cell grade + kelly_advice formulas survive.

Why this matters: the user's original ask was "I want a formula inside
a cell that grades the pick D thru Aplus. Another cell that gives Kelly
advice." If anything breaks the formula-emission path (shape change,
column reorder, regex hit) one of these tests will fail loud.
"""
from __future__ import annotations

import tempfile
from pathlib import Path

import pytest
from openpyxl import load_workbook

from edge_equation.exporters.mlb.daily_spreadsheet import DailySpreadsheet


@pytest.fixture
def synthetic_data():
    """Minimal projection-only data dict matching DailySpreadsheet.collect()
    output. backfill rows + per-tab projection_columns including the
    edge_pct + kelly_pct fields the formulas reference."""
    proj_columns = [
        "matchup", "pick", "model_prob",
        "edge_pct", "kelly_pct", "kelly_advice",
    ]
    backfill_columns = ["date", "actual"]
    return {
        "today": "2026-05-04",
        "season": 2026,
        "counts": {"backfill_games": 100},
        "tabs": {
            "todays_card": {
                "title": "Today's Card",
                "projection_columns": proj_columns,
                "projections": [
                    {"matchup": "AZ@CHC", "pick": "CHC", "model_prob": 0.62,
                     "edge_pct": 9.5, "kelly_pct": 4.2, "kelly_advice": "2u"},
                    {"matchup": "NYY@BOS", "pick": "NYY", "model_prob": 0.55,
                     "edge_pct": 3.5, "kelly_pct": 1.8, "kelly_advice": "0.5u"},
                    {"matchup": "LAD@SF", "pick": "LAD", "model_prob": 0.51,
                     "edge_pct": 0.5, "kelly_pct": 0.3, "kelly_advice": "PASS"},
                    {"matchup": "MIA@WSH", "pick": "MIA", "model_prob": 0.49,
                     "edge_pct": -2.5, "kelly_pct": 0.0, "kelly_advice": "PASS"},
                ],
                "backfill_columns": backfill_columns,
                "backfill": [],
            },
            "moneyline":     _empty_tab("Moneyline", proj_columns, backfill_columns),
            "run_line":      _empty_tab("Run Line", proj_columns, backfill_columns),
            "totals":        _empty_tab("Totals", proj_columns, backfill_columns),
            "first_5":       _empty_tab("First 5", proj_columns, backfill_columns),
            "first_inning":  _empty_tab("First Inning", proj_columns, backfill_columns),
            "team_totals":   _empty_tab("Team Totals", proj_columns, backfill_columns),
            "backtest":      _empty_tab(
                "Backtest",
                ["scope", "bet_type", "bets", "wins", "losses", "pushes",
                 "hit_rate", "units_pl", "roi_pct", "brier"],
                ["date", "daily_units", "cumulative_units"],
            ),
        },
    }


def _empty_tab(title: str, proj_cols: list[str], bf_cols: list[str]) -> dict:
    return {
        "title": title,
        "projection_columns": proj_cols,
        "projections": [],
        "backfill_columns": bf_cols,
        "backfill": [],
    }


def _build_xlsx(data: dict) -> Path:
    inst = DailySpreadsheet.__new__(DailySpreadsheet)
    tmp_dir = Path(tempfile.mkdtemp())
    out = tmp_dir / "mlb_daily.xlsx"
    inst._write_xlsx(out, data)
    return out


def test_grade_column_appended_to_xlsx_view_only(synthetic_data):
    """The 'grade' column is added to xlsx_cols when edge_pct exists,
    but the underlying tab's projection_columns list is unchanged so
    CSV / JSON outputs aren't affected."""
    original_cols = list(synthetic_data["tabs"]["todays_card"]["projection_columns"])
    out = _build_xlsx(synthetic_data)
    # Source data unchanged
    assert synthetic_data["tabs"]["todays_card"]["projection_columns"] == original_cols
    # xlsx headers include 'grade' at the end
    wb = load_workbook(out)
    ws = wb["Today's Card"]
    headers = [ws.cell(row=2, column=c).value for c in range(1, 8)]
    assert headers == [
        "matchup", "pick", "model_prob",
        "edge_pct", "kelly_pct", "kelly_advice", "grade",
    ]


def test_grade_cell_is_formula_referencing_edge_pct(synthetic_data):
    out = _build_xlsx(synthetic_data)
    wb = load_workbook(out)
    ws = wb["Today's Card"]
    # Grade is in column G (7th, after matchup/pick/model_prob/edge_pct/kelly_pct/kelly_advice)
    grade_cell = ws["G3"].value
    assert isinstance(grade_cell, str)
    assert grade_cell.startswith("=IF("), grade_cell
    # The formula must reference column D (edge_pct), not the value
    assert "D3" in grade_cell
    # And cover the full grade ladder
    for letter in ("A+", "A", "B", "C", "D", "F"):
        assert f'"{letter}"' in grade_cell


def test_kelly_advice_cell_is_formula_referencing_kelly_pct(synthetic_data):
    out = _build_xlsx(synthetic_data)
    wb = load_workbook(out)
    ws = wb["Today's Card"]
    # Kelly advice is column F
    cell = ws["F3"].value
    assert cell.startswith("=IF("), cell
    assert "E3" in cell  # references kelly_pct column
    for tier in ("PASS", "0.5u", "1u", "2u", "3u"):
        assert f'"{tier}"' in cell


def test_grade_formula_per_row_uses_correct_row_index(synthetic_data):
    out = _build_xlsx(synthetic_data)
    wb = load_workbook(out)
    ws = wb["Today's Card"]
    # 4 projection rows → rows 3, 4, 5, 6
    for offset, expected_row in enumerate([3, 4, 5, 6]):
        cell = ws.cell(row=expected_row, column=7).value
        assert f"D{expected_row}" in cell, (expected_row, cell)
        # Kelly column too
        kelly = ws.cell(row=expected_row, column=6).value
        assert f"E{expected_row}" in kelly


def test_grade_thresholds_match_scoring_module(synthetic_data):
    """Pin the grade ladder to scoring.py's thresholds. If those drift,
    the workbook's formula must follow."""
    out = _build_xlsx(synthetic_data)
    wb = load_workbook(out)
    grade_formula = wb["Today's Card"]["G3"].value
    # Per scoring.py: A+>=0.08, A>=0.05, B>=0.03, C>=0.00, D>=-0.03 (in fraction)
    # Workbook stores edge as percent so thresholds become 8/5/3/0/-3.
    assert ">=8" in grade_formula
    assert ">=5" in grade_formula
    assert ">=3" in grade_formula
    assert ">=0" in grade_formula
    assert ">=-3" in grade_formula


def test_brand_metadata_stamped_in_workbook_properties(synthetic_data):
    out = _build_xlsx(synthetic_data)
    wb = load_workbook(out)
    # Subject carries the BRAND_TAGLINE constant.
    assert wb.properties.subject == "Facts. Not Feelings."
    assert "Edge Equation" in (wb.properties.title or "")
    assert "Edge Equation" in (wb.properties.creator or "")


def test_empty_projections_does_not_emit_grade_formula(synthetic_data):
    """Empty market tabs should still have headers (incl. grade) but no
    grade formulas in non-existent rows."""
    out = _build_xlsx(synthetic_data)
    wb = load_workbook(out)
    ws = wb["Moneyline"]  # synthetic_data has 0 projection rows here
    headers = [ws.cell(row=2, column=c).value for c in range(1, 8)]
    assert "grade" in headers
    # Row 3 is empty (or starts the backfill section). No formula expected.
    g3 = ws.cell(row=3, column=7).value
    assert g3 is None or not str(g3).startswith("=IF(")


def test_formula_recomputes_when_edge_pct_changes_in_libreoffice():
    """Sanity: openpyxl preserves the formula as a string, not a value.
    A consumer (Excel / LibreOffice / Numbers) opening the file must see
    the formula and recompute when underlying cells change. We can't
    invoke a calc engine in pytest, but we can prove openpyxl wrote the
    formula via data_only=False (default) and didn't pre-evaluate."""
    proj_columns = [
        "matchup", "pick", "model_prob", "edge_pct", "kelly_pct", "kelly_advice",
    ]
    data = {
        "today": "2026-05-04", "season": 2026,
        "counts": {"backfill_games": 0},
        "tabs": {
            "todays_card": {
                "title": "Today's Card",
                "projection_columns": proj_columns,
                "projections": [
                    {"matchup": "X@Y", "pick": "X", "model_prob": 0.6,
                     "edge_pct": 7.5, "kelly_pct": 3.0, "kelly_advice": "1u"},
                ],
                "backfill_columns": ["date"], "backfill": [],
            },
            "moneyline": _empty_tab("Moneyline", proj_columns, ["date"]),
            "run_line": _empty_tab("Run Line", proj_columns, ["date"]),
            "totals": _empty_tab("Totals", proj_columns, ["date"]),
            "first_5": _empty_tab("First 5", proj_columns, ["date"]),
            "first_inning": _empty_tab("First Inning", proj_columns, ["date"]),
            "team_totals": _empty_tab("Team Totals", proj_columns, ["date"]),
            "backtest": _empty_tab("Backtest", ["scope"], ["date"]),
        },
    }
    out = _build_xlsx(data)
    # Default load: read formulas as strings.
    wb_formulas = load_workbook(out, data_only=False)
    g = wb_formulas["Today's Card"]["G3"].value
    assert g is not None and g.startswith("=IF(")
    # data_only=True would normally read the cached calc result — but
    # openpyxl doesn't compute formulas, so cached_value is None until
    # a calc engine touches the file. Accepting None confirms openpyxl
    # is NOT pre-evaluating and is leaving recomputation to the consumer.
    wb_values = load_workbook(out, data_only=True)
    cached = wb_values["Today's Card"]["G3"].value
    assert cached is None  # recomputation deferred to Excel/Numbers/LibreOffice
