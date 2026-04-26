"""
Generates `edge_equation_props.xlsx` -- a Numbers-compatible starter
workbook for manual prop-tracking math on iPhone / iPad.

Numbers iOS opens .xlsx natively and converts it to .numbers on save.
Formulas in this workbook use the .xlsx Sheet!Range reference style
which Numbers translates to its Table::Range style on import. The
defined name `payout_mult` survives the import as a workbook-scoped
named reference.

Run:
    python3 tools/sheets/build_numbers_workbook.py

Output:
    tools/sheets/edge_equation_props.xlsx

Workbook layout matches numbers_shortcut_runbook.md:

  Sheet `_raw`   : 12-column landing zone for the Shortcut's clipboard
                   paste. Header row only -- gets overwritten on every
                   refresh.

  Sheet `picks`  : 13-column handicapping worksheet. Row 2 carries the
                   formulas (auto-fill via VLOOKUP into _raw, then the
                   break_even / edge / half_kelly / grade chain). The
                   formulas are pre-dragged through row 50 so any future
                   projection_id pasted into column A works without
                   editing.

                   Cell P1 / Q1 hold the `payout_mult` setting (default
                   3 for a 2-pick power play). Defined name `payout_mult`
                   points to Q1, so all formulas just reference it by
                   name; change Q1 once and every formula recomputes.

  Sheet `tracking` : 10-column results log. Hit / units P&L formulas
                     pre-filled through row 50.

The math here mirrors src/edge_equation/math/ev.py (break_even = sqrt
(1/payout), edge = your_prob - break_even, half_kelly capped at 25%)
and src/edge_equation/math/scoring.py (grade thresholds A+ >= 0.10,
A >= 0.07, B >= 0.04, C >= 0.02, D > 0, else F). If those constants
ever drift, this script's `RAW_HEADERS` / formula strings need to
follow.
"""
from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName


# Sheet names
RAW = "_raw"
PICKS = "picks"
TRACKING = "tracking"

RAW_HEADERS = [
    "scraped_at", "projection_id", "league", "sport", "player",
    "team", "position", "description", "stat_type", "line",
    "start_time", "odds_type",
]
PICKS_HEADERS = [
    "projection_id", "player", "team", "league", "stat_type",
    "line", "pick (over/under)", "your_prob", "break_even",
    "edge", "half_kelly", "grade", "notes",
]
TRACKING_HEADERS = [
    "projection_id", "player", "stat_type", "line", "pick",
    "actual", "hit", "units_risked", "pl_units", "note",
]

# Cyan accent + ink charcoal -- matches the brand tokens in
# website/tailwind.config.js so the workbook reads consistent with
# the Edge Equation visual language.
HEADER_FILL = PatternFill("solid", fgColor="1E2229")  # ink-700
HEADER_FONT = Font(name="Helvetica Neue", size=12, bold=True, color="5BC0E8")
DATA_FONT = Font(name="Helvetica Neue", size=11)


def _write_header(ws, headers):
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.freeze_panes = "A2"


def _set_column_widths(ws, widths):
    for col_idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def build() -> Workbook:
    wb = Workbook()

    # -------------------------------------------------- _raw
    raw = wb.active
    raw.title = RAW
    _write_header(raw, RAW_HEADERS)
    _set_column_widths(raw, [
        20, 14, 8, 10, 22, 8, 6, 32, 14, 8, 22, 12,
    ])

    # -------------------------------------------------- picks
    picks = wb.create_sheet(PICKS)
    _write_header(picks, PICKS_HEADERS)
    _set_column_widths(picks, [
        14, 22, 8, 8, 14, 8, 14, 12, 12, 12, 12, 8, 30,
    ])

    # Settings: P1 holds the label, Q1 the value. Defined name
    # `payout_mult` points at Q1 so formulas can reference it by
    # name (and Numbers preserves the named ref on .xlsx import).
    picks["P1"] = "payout_mult ↓"
    picks["P1"].font = Font(italic=True, color="8A8A7A")
    picks["Q1"] = 3
    picks["Q1"].font = Font(bold=True, size=14, color="5BC0E8")
    picks["Q1"].alignment = Alignment(horizontal="center")

    picks["P2"] = "2-pick=3 · 3-pick=5 · 4-pick=10 · 5-pick=20 · 6-pick=25"
    picks["P2"].font = Font(italic=True, color="8A8A7A", size=9)

    # The formulas. Built once for row 2, then duplicated through row 50.
    # break_even uses payout_mult directly; edge / half_kelly chain off it.
    formula_templates = {
        "B": '=IFERROR(VLOOKUP(A{row},_raw!B:L,4,FALSE),"")',
        "C": '=IFERROR(VLOOKUP(A{row},_raw!B:L,5,FALSE),"")',
        "D": '=IFERROR(VLOOKUP(A{row},_raw!B:L,2,FALSE),"")',
        "E": '=IFERROR(VLOOKUP(A{row},_raw!B:L,8,FALSE),"")',
        "F": '=IFERROR(VLOOKUP(A{row},_raw!B:L,9,FALSE),"")',
        "I": '=IF(payout_mult="","",SQRT(1/payout_mult))',
        "J": '=IF(OR(H{row}="",I{row}=""),"",H{row}-I{row})',
        "K": (
            '=IFERROR(IF(OR(H{row}="",J{row}<=0),0,'
            'MIN(0.25,MAX(0,(H{row}*(payout_mult^(1/2)-1)-(1-H{row}))/'
            '(payout_mult^(1/2)-1)/2))),0)'
        ),
        "L": (
            '=IF(J{row}="","",'
            'IF(J{row}>=0.10,"A+",'
            'IF(J{row}>=0.07,"A",'
            'IF(J{row}>=0.04,"B",'
            'IF(J{row}>=0.02,"C",'
            'IF(J{row}>0,"D","F"))))))'
        ),
    }
    for row in range(2, 51):
        for col, tmpl in formula_templates.items():
            picks[f"{col}{row}"] = tmpl.format(row=row)
            picks[f"{col}{row}"].font = DATA_FONT

    # Number formats. .xlsx percent format codes survive the
    # Numbers import.
    for row in range(2, 51):
        picks[f"F{row}"].number_format = "0.0"          # line: one decimal
        picks[f"H{row}"].number_format = "0.00%"        # your_prob
        picks[f"I{row}"].number_format = "0.00%"        # break_even
        picks[f"J{row}"].number_format = "0.00%"        # edge
        picks[f"K{row}"].number_format = "0.00%"        # half_kelly
        picks[f"L{row}"].alignment = Alignment(horizontal="center")

    # -------------------------------------------------- tracking
    tracking = wb.create_sheet(TRACKING)
    _write_header(tracking, TRACKING_HEADERS)
    _set_column_widths(tracking, [
        14, 22, 14, 8, 8, 8, 8, 12, 12, 30,
    ])
    for row in range(2, 51):
        tracking[f"G{row}"] = (
            f'=IF(F{row}="","",'
            f'IF(E{row}="over",'
            f'IF(F{row}>D{row},1,IF(F{row}<D{row},0,"push")),'
            f'IF(F{row}<D{row},1,IF(F{row}>D{row},0,"push"))))'
        )
        tracking[f"I{row}"] = (
            f'=IF(G{row}="push",0,'
            f'IF(G{row}=1,H{row}*(picks!Q1^(1/2)-1),'
            f'IF(G{row}=0,-H{row},"")))'
        )
        tracking[f"H{row}"].number_format = "0.00"
        tracking[f"I{row}"].number_format = "0.00"

    # -------------------------------------------------- defined name
    # Workbook-scoped name -- Numbers iOS preserves this on import.
    wb.defined_names["payout_mult"] = DefinedName(
        "payout_mult", attr_text=f"{PICKS}!$Q$1"
    )

    return wb


def main() -> None:
    out = Path(__file__).parent / "edge_equation_props.xlsx"
    wb = build()
    wb.save(out)
    print(f"wrote {out}  ({out.stat().st_size} bytes)")


if __name__ == "__main__":
    main()
